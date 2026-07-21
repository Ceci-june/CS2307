import json
import re
import argparse
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "data" / "exceltemplate.xlsx"
JSON_PATH = BASE_DIR / "data" / "hcm_can_ho_full_detail.json"
OUTPUT_PATH = BASE_DIR / "data" / "hcm_can_ho_full_detail_from_template.xlsx"


def clean_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_feature_obj(row: Dict[str, Any]) -> Dict[str, Any]:
    for key in ("Đặc điểm bất động sản (json)", "Đặc điểm bất động sản"):
        raw = row.get(key)
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, str) and raw.strip():
            try:
                obj = json.loads(raw)
                if isinstance(obj, dict):
                    return obj
            except Exception:
                pass
    return {}


def strip_parentheses(text: Any) -> str:
    if text is None:
        return ""
    value = str(text)
    value = value.replace("(", "").replace(")", "")
    return value.strip()


def format_latlon(row: Dict[str, Any]) -> str:
    lat = str(row.get("Latitude", "") or "").strip()
    lon = str(row.get("Longitude", "") or "").strip()
    if lat and lon:
        return f"q={lat},{lon}"

    link = str(row.get("Link Location", "") or "").strip()
    match = re.search(r"[?&]q=([^&]+)", link)
    if match:
        return f"q={match.group(1)}"

    return ""


def extract_number(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).strip()
    if not value:
        return ""
    match = re.search(r"\d+", value)
    return match.group(0) if match else value


def value_for_col(col: str, row: Dict[str, Any], index: int) -> Any:
    feature = parse_feature_obj(row)

    mapping = {
        "Mã tin": row.get("Mã tin", ""),
        "Loại tin": row.get("Loại tin", ""),
        "Ngày đăng": row.get("Ngày đăng", ""),
        "Ngày hết hạn": row.get("Ngày hết hạn", ""),
        "Loại BĐS": row.get("Loại BĐS", ""),
        "Tỉnh thành": row.get("Tỉnh thành", ""),
        "Quận huyện": row.get("Quận huyện", ""),
        "Folder": row.get("breadcrumb_folder", "") or row.get("Folder", ""),
        "Tiêu đề": row.get("Tiêu đề", ""),
        "Địa chỉ": row.get("Địa chỉ", ""),
        "Địa chỉ cũ": strip_parentheses(row.get("Địa chỉ trước sát nhập", "")),
        "Link Location": row.get("Link Location", ""),
        "Latitude and Longitude": format_latlon(row),
        "Mô tả": row.get("Thông tin mô tả", ""),
        "Lịch sử giá khu": row.get("price_trend", ""),
        "Khoảng giá": row.get("Khoảng giá", "") or row.get("Giá", ""),
        "Diện tích": row.get("Diện tích", ""),
        "Đơn giá": row.get("Đơn giá/m²", ""),
        "Số phòng ngủ": extract_number(
            row.get("Đặc điểm - So Phong Ngu", "")
            or row.get("Phòng ngủ", "")
            or feature.get("Số phòng ngủ", "")
        ),
        "Số phòng tắm, vệ sinh": extract_number(
            row.get("Đặc điểm - So Phong Tam Ve Sinh", "")
            or feature.get("Số phòng tắm, vệ sinh", "")
        ),
        "Pháp lý": row.get("Đặc điểm - Phap Ly", "") or feature.get("Pháp lý", ""),
        "Nội thất": row.get("Đặc điểm - Noi That", "") or feature.get("Nội thất", ""),
        "Hướng ban công": row.get("Đặc điểm - Huong Ban Cong", "") or feature.get("Hướng ban công", ""),
    }

    if col == "#":
        return index

    if col in mapping:
        return mapping[col]

    if col in feature:
        return feature.get(col, "")

    return row.get(col, "")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export JSON listings to Excel template format")
    parser.add_argument("--input", default=str(JSON_PATH), help="Input JSON file path")
    parser.add_argument("--output", default=str(OUTPUT_PATH), help="Output Excel file path")
    parser.add_argument("--template", default=str(TEMPLATE_PATH), help="Template Excel file path")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    template_path = Path(args.template)

    wb_template = load_workbook(template_path, data_only=True)
    ws_template = wb_template[wb_template.sheetnames[0]]

    template_headers_raw = list(
        next(ws_template.iter_rows(min_row=1, max_row=1, values_only=True))
    )
    template_headers = [clean_header(h) for h in template_headers_raw if clean_header(h)]

    rows: List[Dict[str, Any]] = json.loads(input_path.read_text(encoding="utf-8"))

    # Add extra columns from nested property features only if not in template
    template_lower = {h.lower() for h in template_headers}
    extra_feature_cols: List[str] = []
    seen_extra = set()

    for row in rows:
        feature = parse_feature_obj(row)
        for key in feature.keys():
            col_name = str(key).strip()
            if not col_name:
                continue
            lower_name = col_name.lower()
            if lower_name not in template_lower and lower_name not in seen_extra:
                seen_extra.add(lower_name)
                extra_feature_cols.append(col_name)

    final_headers = template_headers + extra_feature_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws.append(final_headers)
    for index, row in enumerate(rows, start=1):
        ws.append([value_for_col(col, row, index) for col in final_headers])

    wb.save(output_path)

    print(f"Created: {output_path}")
    print(f"Rows: {len(rows)}")
    print(f"Columns: {len(final_headers)}")
    print(f"Extra feature columns: {extra_feature_cols}")


if __name__ == "__main__":
    main()
